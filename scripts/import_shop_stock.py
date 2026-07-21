#!/usr/bin/env python3
"""
Import STOCK Excel → src/data/shop-stock.json (sell price + qty only).

POS (ProjectOmahBan) owns live stock. This is a read-only snapshot for
Wheelpedia counter recommendations. NEVER write cost/modal to JSON.

Usage:
    python scripts/import_shop_stock.py "C:\\Users\\cthrn\\Downloads\\STOCK JULI 2026.xlsx"
    npm run import-stock -- "C:\\path\\to\\STOCK.xlsx"

Deps: openpyxl only
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = ROOT / "src" / "data" / "shop-stock.json"

# Sheet → canonical brand (None = parse from product name)
SHEET_BRAND_MAP = {
    "bridgestone": "Bridgestone",
    "dunlop": "Dunlop",
    "gt ( gajah tunggal )": "GT",
    "acellera": "Accelera",
    "delium": "Delium",
    "campur": None,
    "truck diesel": None,
    "ban dalam": None,
}

BRAND_PREFIXES = {
    "bs ": "Bridgestone",
    "bridgestone": "Bridgestone",
    "gt ": "GT",
    "gajah tunggal": "GT",
    "acc ": "Accelera",
    "accelera": "Accelera",
    "acellera": "Accelera",
    "dunlop": "Dunlop",
    "hankook": "Hankook",
    "goodyear": "Goodyear",
    "sailun": "Sailun",
    "achilles": "Achilles",
    "swallow": "Swallow",
    "pirelli": "Pirelli",
    "continental": "Continental",
    "yokohama": "Yokohama",
    "delium": "Delium",
    "forceum": "Forceum",
    "aeolus": "Aeolus",
    "falken": "Falken",
    "federal": "Federal",
}


def cell_val(row, idx: int):
    """1-based column index from openpyxl row tuple (0-based list)."""
    if idx - 1 >= len(row):
        return None
    return row[idx - 1]


def clean_rupiah(val) -> int | None:
    if val is None:
        return None
    try:
        if isinstance(val, str):
            val = val.replace(".", "").replace(",", "").strip()
            if not val:
                return None
        return int(float(val))
    except (ValueError, TypeError):
        return None


def clean_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def extract_year_hint(product_name: str) -> tuple[str | None, str]:
    if not product_name:
        return (None, product_name or "")
    pattern = re.compile(
        r"\((?:th\s*)?(\d{2})(?:[\s,]*\d{2})*(?:\s*\+\s*\d{2})?\)",
        re.IGNORECASE,
    )
    match = pattern.search(product_name)
    year_hint = None
    if match:
        digits = re.findall(r"\d{2}", match.group(0))
        if digits:
            max_digit = max(int(d) for d in digits)
            if 20 <= max_digit <= 29:
                year_hint = str(2000 + max_digit)
            elif 90 <= max_digit <= 99:
                year_hint = str(1900 + max_digit)
    cleaned = pattern.sub("", product_name).strip()
    cleaned = re.sub(r"@[\d.]+", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return (year_hint, cleaned)


def extract_brand_from_name(name: str) -> str | None:
    if not name:
        return None
    lower = name.lower().strip()
    for prefix in sorted(BRAND_PREFIXES.keys(), key=len, reverse=True):
        if lower.startswith(prefix):
            return BRAND_PREFIXES[prefix]
    first = lower.split()[0] if lower.split() else ""
    return BRAND_PREFIXES.get(first) or (first.title() if first else None)


def resolve_brand(sheet_name: str, raw_name: str, clean_name: str) -> str | None:
    key = sheet_name.lower().strip()
    if key in SHEET_BRAND_MAP and SHEET_BRAND_MAP[key] is not None:
        return SHEET_BRAND_MAP[key]
    if key == "campur":
        return extract_brand_from_name(clean_name or raw_name)
    from_raw = extract_brand_from_name(raw_name)
    if from_raw:
        return from_raw
    return extract_brand_from_name(clean_name) or sheet_name.strip()


def normalize_tire_size(size_raw: str, rim) -> str:
    """Keep in sync with scripts/lib/tire-size.mjs / src/lib/shop-stock.ts"""
    if size_raw is None or rim is None:
        return ""
    try:
        rim_n = int(float(rim))
    except (ValueError, TypeError):
        return ""
    s = str(size_raw).strip().replace(" ", "")
    if not s or s.lower() == "nan":
        return ""
    if re.fullmatch(r"\d{3}/\d{2}", s):
        return f"{s} R{rim_n}"
    if re.fullmatch(r"\d{2,3}", s):
        return f"{s} R{rim_n}"
    if re.search(r"R\d{2}$", s, re.IGNORECASE):
        out = re.sub(r"r", " R", s, count=1, flags=re.IGNORECASE)
        out = re.sub(r"\s+", " ", out).upper().strip()
        out = re.sub(r"\s+R", " R", out)
        return out
    return f"{s} R{rim_n}"


def import_excel(excel_path: Path) -> list[dict]:
    print(f"Reading: {excel_path}")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    items: list[dict] = []
    idx = 0
    skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Data starts row 5; cols: 1=no 2=merk 3=ukuran 4=ring 5=modal 6=jual 7=stock 8=sisa
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row:
                skipped += 1
                continue
            raw_cell = cell_val(row, 2)
            raw_name = str(raw_cell).strip() if raw_cell is not None else ""
            if not raw_name or raw_name.startswith("*") or "ambil" in raw_name.lower():
                skipped += 1
                continue

            size_cell = cell_val(row, 3)
            size_raw = str(size_cell).strip() if size_cell is not None else ""
            if size_raw.lower() == "nan":
                size_raw = ""

            ring_cell = cell_val(row, 4)
            rim: int | None
            try:
                rim = int(float(ring_cell)) if ring_cell is not None else None
            except (ValueError, TypeError):
                rim = None

            year_hint, product_name = extract_year_hint(raw_name)
            if not product_name:
                skipped += 1
                continue
            if not size_raw or rim is None:
                skipped += 1
                continue

            brand = resolve_brand(sheet_name, raw_name, product_name)
            if not brand:
                skipped += 1
                continue

            sell = clean_rupiah(cell_val(row, 6))
            if sell is None:
                sell = 0

            qty = clean_int(cell_val(row, 8))  # Sisa
            if qty is None:
                qty = clean_int(cell_val(row, 7))  # Stock fallback
            if qty is None:
                qty = 0

            rim_n = rim if rim is not None else 0
            size_norm = normalize_tire_size(size_raw, rim_n) if size_raw else ""

            idx += 1
            item = {
                "id": f"ob-{idx:05d}",
                "brand": brand,
                "productName": product_name,
                "sizeRaw": size_raw,
                "rim": rim_n,
                "sizeNormalized": size_norm,
                "sellPrice": sell,
                "qty": qty,
                "sheet": sheet_name,
            }
            if year_hint:
                item["yearHint"] = year_hint
            items.append(item)

    wb.close()
    print(f"Imported {len(items)} rows (skipped {skipped})")
    return items


def main() -> int:
    if len(sys.argv) < 2:
        print(
            "Usage: python scripts/import_shop_stock.py <path-to-STOCK.xlsx> [out.json]",
            file=sys.stderr,
        )
        return 2

    excel = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not excel.is_file():
        print(f"File not found: {excel}", file=sys.stderr)
        return 1

    items = import_excel(excel)
    if not items:
        print("No rows imported — abort write", file=sys.stderr)
        return 1

    for it in items:
        for bad in ("costPrice", "modal", "product_cost", "avg_cost", "batch_cost"):
            it.pop(bad, None)

    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"Wrote {len(items)} items -> {out}")
    sample = items[0]
    print("Sample keys:", sorted(sample.keys()))
    print("Sample:", sample)
    assert "costPrice" not in sample and "modal" not in sample
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
